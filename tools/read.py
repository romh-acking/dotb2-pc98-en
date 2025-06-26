# pip3 install openpyxl

from openpyxl import load_workbook
from enum import Enum
import re

class Columns(Enum):
    File = 1
    Line = 2
    Marker = 3
    Original = 4
    Speaker	= 5
    HumanReadableLine = 6	
    Speaker_EN = 7
    HumanReadableLine_EN = 8

fileName = 'Copy of DOTB2 Worksheet_v4_Edits.xlsx'

wb = load_workbook(fileName)
sheet = wb.active 

max_row = sheet.max_row

replaceCache = {}

def getText(column):
    return sheet.cell(row=row, column=column.value).value

for row in range(2, max_row + 1):
    file = getText(Columns.File)
    if not file:
        continue

    file = file.upper()
    file = file.replace(".RKT", ".rkt")

    findText = getText(Columns.Original)

    marker = getText(Columns.Marker)
    speaker = getText(Columns.Speaker_EN)
    line = getText(Columns.HumanReadableLine_EN)

    if not line:
        continue

    if speaker:
        speaker = f"{speaker}: "
    else:
        speaker = ""

    replaceText = f"{speaker}{line}"

    # Check if line doesn't already do word wrapping
    if "  "  not in replaceText:
        split = replaceText.split(" ")

        lineCount = 0
        replaceText = ""
        charaWidth = 60

        #linebreak = "".rjust(slack)
        linebreak = f"\" 'br)\n{marker}"

        for x in split:
            # Right on the edge
            if lineCount == charaWidth:
                replaceText = f"{replaceText}{linebreak}{x}"
                lineCount = len(x)
                continue

            # Add character on same line
            if lineCount + len(x) + 1 <= charaWidth:

                if replaceText != "":
                    replaceText = f"{replaceText} {x}"
                    lineCount += len(x) + 1
                else:
                    replaceText = f"{x}"
                    lineCount += len(x)

                continue

            # Line break
            slack = charaWidth - lineCount
            replaceText = f"{replaceText}{linebreak}{x}"


            lineCount = len(x)
    finalReplaceText = re.sub(r'["\'](.*?)["\']', f'"{replaceText}"', findText)

    if file not in replaceCache:
        replaceCache[file] = []    
    replaceCache[file].append({"find": findText, "replace": finalReplaceText})
    print()

for fileKey in replaceCache:
    path = f'./script/{fileKey}'

    with open(path, 'r') as file:
        filedata = file.read()

    filedata = filedata.replace("(charset \"pc98\")", "(charset \"English\")")

    for r in replaceCache[fileKey]:
        filedata = filedata.replace(r["find"], r["replace"])

    with open(path, 'w') as file:
        file.write(filedata)


    print()